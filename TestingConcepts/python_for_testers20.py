"""
Test Data Generation: Using faker module
1. Helps to generate fake data
"""
import faker
from faker.proxy import Faker
from openpyxl import Workbook

path = "C:\\Users\\ashok\\PycharmProjects\\cloned_project\\my_python_selenium_pytest_git_jenkins_learnings\\TestingConcepts\\Files\\"
wb = Workbook()
ws = wb.active
# fake_data = Faker()
fake_data = Faker(['hi_IN','en_US'])    # If you want data in any specific language
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())

for i in range(1, 11):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()
    ws.cell(row=i, column=3).value = fake_data.address()
wb.save(path+"testData.xlsx")