"""
Using openpyxl 3.1.5
Writing in XLSX files.
"""
from openpyxl import Workbook

path = "C:\\Users\\ashok\\PycharmProjects\\cloned_project\\my_python_selenium_pytest_git_jenkins_learnings\\TestingConcepts\\Files\\"
wb = Workbook() #To open workbook
ws = wb.active # grab the active worksheet
# ws["A1"] = "Ashok Kumar"    # Data can be assigned directly to cells
# ws["C3"] = "C column"
# ws.append(["Ashok", "Bengaluru"])
# ws.append(["Kumar", "Reddy"])
testdata = [["Name", "City"],["Ashok", "Bengaluru"],["Suri", "Marthahalli"], ["Krishna", "Chinnapannahalli"]]
for data in testdata:
    print(data)
    ws.append(data)

for i in range(1,6):
    for j in range(1,5):
        ws.cell(row=i+4, column=j).value = i+j
ws = wb.create_sheet("Report")
wb["Report"]["A1"] = "Summary"
wb.save(path+"sample.xlsx")# Save the file