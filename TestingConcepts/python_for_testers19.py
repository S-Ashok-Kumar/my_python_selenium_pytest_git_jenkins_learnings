"""
Using openpyxl 3.1.5
Reading from XLSX files.
"""
from openpyxl import  Workbook, load_workbook

path = "C:\\Users\\ashok\\PycharmProjects\\cloned_project\\my_python_selenium_pytest_git_jenkins_learnings\\TestingConcepts\\Files\\"
wb = load_workbook(path+"sample.xlsx")
# sh = wb.active # grab the active worksheet
sh = wb["Sheet1"]
print(sh["A3"].value)   # 1st way to access data from sheet
# print(wb["Sheet1"]['A3'].value) # or 2nd way to access data from sheet
# print(sh.cell(2,2).value)   # or 3rd way to access data from sheet

row_count = sh.max_row  # to know how many rows in sheet
column_count = sh.max_column    # to know how many columns in sheet
print(row_count, column_count)

for i in range(1,row_count+1):
    for j in range(1,column_count+1):
        print(sh.cell(row=i,column=j).value,end='')
    print("\n")
